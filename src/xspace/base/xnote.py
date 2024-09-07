import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from collections import OrderedDict
from openpyxl.drawing.image import Image
from loguru import logger
from typing import Any, Sequence, Dict, List, Tuple

class XNoteRow:
    def __init__(self, row: Sequence):
        self._row = row
    @property 
    def row(self):
        return self._row
    
    def __len__(self):
        return len(self._row)
    
    def __getitem__(self, key):
        return self._row[key]

class XNoteCol:
    def __init__(self, col: Sequence):
        self._col = col
        
    @property 
    def col(self):
        return self._col
    
    def __len__(self):
        return len(self._col)
    
    def __getitem__(self, key):
        return self._col[key]
 
class XNote:
    def __init__(self, xlsx_file, sheet_name='Sheet1', image_width=256, mode='w'):
        self.xlsx_file = Path(xlsx_file)
        if self.xlsx_file.exists():
            self.work_book = load_workbook(self.xlsx_file)
            sheet_names = self.work_book.sheetnames
            isname = sheet_name
            if mode == 'w':
                if sheet_name in sheet_names:
                    self.work_book.remove(self.work_book[sheet_name])
                    logger.warning(f"Sheet {sheet_name} already exists, remove it.")
            else:
                num = 0
                while isname in sheet_names:
                    isname = f"{sheet_name}_{num+1}"
                    num += 1
                if num > 0:
                    logger.warning(f"Sheet name {sheet_name} already exists, use {isname} instead.")
            self.work_book_sheet = self.work_book.create_sheet(title=isname)
        else:
            self.work_book = Workbook()
            self.work_book_sheet = self.work_book.active

        self.image_width = image_width
    
    @classmethod 
    def from_dict(cls, xlsx_file, data: Dict, sheet_name='Sheet1', image_width=256, mode='w'):
        obj = cls(xlsx_file, sheet_name, image_width, mode)

        for idx, key in enumerate(data.keys()):
            obj[idx] = XNoteCol([key] + data[key])

        return obj
    
    @classmethod
    def from_list(cls, xlsx_file, data: List[Dict|List|Tuple], head: List[str], sheet_name='Sheet1', image_width=256, mode='w'):
        obj = cls(xlsx_file, sheet_name, image_width, mode)

        obj[0] = XNoteRow(head)
        nhead = len(head)
        if isinstance(data[0], (List, Tuple)):
            for row, klist in enumerate(data):
                row += 1
                obj[row] = XNoteRow(klist[:nhead])
        else:
            for row, kdict in enumerate(data):
                row += 1
                obj[row] = XNoteRow([kdict.get(k, None) for k in head])
        
        return obj
    
    def __resize_image(self, image: Image):
        src_w, src_h = image.width, image.height
        if src_w > src_h:
            dst_w = self.image_width 
            dst_h = int(src_h * self.image_width / src_w)
        else:
            dst_h = self.image_width 
            dst_w = int(src_w * self.image_width / src_h)
        return dst_w, dst_h
    
    def col_to_letter(self, col):
        letters = ''
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def num_to_cell_ref(self, row, col):
        col_letter = self.col_to_letter(col)
        return f"{col_letter}{row}"
    
    def set_img(self, row, col, image):
        dstw, dsth = self.__resize_image(image)
        self.work_book_sheet.column_dimensions[self.col_to_letter(col)].width = dstw * 0.14
        image.width, image.height = dstw, dsth
        self.work_book_sheet.add_image(image, self.num_to_cell_ref(row, col))
        self.work_book_sheet.row_dimensions[row].height = dsth * 0.78 

    def __setitem__(self, key, value: XNoteRow | XNoteCol | Image | Any):
        if isinstance(key, int):
            if isinstance(value, (XNoteRow, XNoteCol)):
                if isinstance(value, XNoteCol):
                    for idx, val in enumerate(value):
                        cell = self.work_book_sheet.cell(row=idx+1, column=key+1)
                        if isinstance(val, Image):
                            self.set_img(idx+1, key+1, val)
                        else:
                            cell.value = val
                elif isinstance(value, XNoteRow):
                    for idx, val in enumerate(value):
                        cell = self.work_book_sheet.cell(row=key+1, column=idx+1)
                        if isinstance(val, Image):
                            self.set_img(key+1, idx+1, val)
                        else:
                            cell.value = val
                else:
                    raise ValueError("Invalid value")
            else:
                cell = self.work_book_sheet.cell(row=1, column=key+1)
                cell.value = value
        elif isinstance(key, tuple) and len(key) == 2:
            y, x = (key[0]+1, key[1]+1)
            if isinstance(x, int) and isinstance(y, int):
                if isinstance(value, Image):
                    self.set_img(y, x, value)
                    # dstw, dsth = self.__resize_image(value)
                    # self.work_book_sheet.column_dimensions[self.col_to_letter(x)].width = dstw * 0.14
                    # value.width, value.height = dstw, dsth
                    # self.work_book_sheet.add_image(value, self.num_to_cell_ref(y, x))
                    # self.work_book_sheet.row_dimensions[y].height = dsth * 0.78
                else:
                    self.work_book_sheet.cell(row=y, column=x, value=value)
            else:
                raise ValueError("Invalid key")
        else:
            raise ValueError("Invalid key")
    
    def __getitem__(self, key):
        if isinstance(key, int):
            vals = []
            for col in self.work_book_sheet.iter_cols(min_col=0, min_row=key, max_row=key+1, values_only=True):
                vals.append(col)
            try:
                return OrderedDict(vals)
            except:
                return [ks[0] for ks in vals]
        elif isinstance(key, tuple) and len(key) == 2:
            x, y = (key[0]+1, key[1]+1)
            if isinstance(x, int) and isinstance(y, int):
                return self.work_book_sheet.cell(row=y, column=x).value
            else:
                raise ValueError("Invalid key")
        else:
            raise ValueError("Invalid key")
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def __del__(self):
        self.save()
    
    def save(self):
        ncol = len([col for col in self.work_book_sheet.columns])
        for col in range(ncol):
            value = self.work_book_sheet.cell(row=1, column=col+1)
            if value is None:
                self.work_book_sheet.cell(row=1, column=col).value = f'column_{col-1}'

        if self.xlsx_file.exists():
           os.remove(self.xlsx_file) 
        self.work_book.save(self.xlsx_file)

